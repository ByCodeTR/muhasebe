"""
Reports and export endpoints.
"""
import io
import uuid
from datetime import datetime, timedelta
from decimal import Decimal

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession
import openpyxl
from openpyxl.utils import get_column_letter

from app.database import get_db
from app.models.ledger import LedgerEntry, EntryDirection
from app.models.vendor import Vendor
# from app.models.category import Category  # Not used in this file
from app.schemas.ledger import ReportSummary

router = APIRouter(prefix="/reports", tags=["Reports"])


@router.get("/summary", response_model=ReportSummary)
async def get_summary(
    period: str = Query("month", regex="^(week|month|year|custom)$"),
    start_date: datetime | None = None,
    end_date: datetime | None = None,
    db: AsyncSession = Depends(get_db),
):
    """Get summary statistics for a period."""
    now = datetime.utcnow()
    
    # Calculate period dates
    if period == "week":
        period_start = now - timedelta(days=7)
        period_end = now
    elif period == "month":
        period_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        period_end = now
    elif period == "year":
        period_start = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
        period_end = now
    else:  # custom
        period_start = start_date or now - timedelta(days=30)
        period_end = end_date or now
    
    # Get totals
    query = select(
        func.sum(LedgerEntry.amount).filter(
            LedgerEntry.direction == EntryDirection.INCOME.value
        ).label("income"),
        func.sum(LedgerEntry.amount).filter(
            LedgerEntry.direction == EntryDirection.EXPENSE.value
        ).label("expense"),
        func.sum(LedgerEntry.tax_amount).label("tax"),
        func.count(LedgerEntry.id).label("count"),
    ).where(
        LedgerEntry.entry_date >= period_start,
        LedgerEntry.entry_date <= period_end,
    )
    
    result = await db.execute(query)
    row = result.one()
    
    income = row.income or Decimal("0")
    expense = row.expense or Decimal("0")
    tax = row.tax or Decimal("0")
    
    return ReportSummary(
        period_start=period_start,
        period_end=period_end,
        total_income=income,
        total_expense=expense,
        net=income - expense,
        tax_total=tax,
        transaction_count=row.count,
    )


@router.get("/by-vendor")
async def get_by_vendor(
    period: str = Query("month"),
    db: AsyncSession = Depends(get_db),
):
    """Get spending breakdown by vendor."""
    now = datetime.utcnow()
    
    if period == "week":
        period_start = now - timedelta(days=7)
    elif period == "month":
        period_start = now.replace(day=1)
    else:
        period_start = now.replace(month=1, day=1)
    
    query = (
        select(
            Vendor.display_name,
            func.sum(LedgerEntry.amount).label("total"),
            func.count(LedgerEntry.id).label("count"),
        )
        .join(LedgerEntry, Vendor.id == LedgerEntry.vendor_id)
        .where(
            LedgerEntry.entry_date >= period_start,
            LedgerEntry.direction == EntryDirection.EXPENSE.value,
        )
        .group_by(Vendor.id, Vendor.display_name)
        .order_by(func.sum(LedgerEntry.amount).desc())
        .limit(10)
    )
    
    result = await db.execute(query)
    
    return [
        {"vendor": row.display_name, "total": float(row.total), "count": row.count}
        for row in result.all()
    ]


@router.get("/export/csv")
async def export_csv(
    start_date: datetime | None = None,
    end_date: datetime | None = None,
    db: AsyncSession = Depends(get_db),
):
    """Export ledger entries as CSV."""
    query = select(LedgerEntry).order_by(LedgerEntry.entry_date.desc())
    
    if start_date:
        query = query.where(LedgerEntry.entry_date >= start_date)
    if end_date:
        query = query.where(LedgerEntry.entry_date <= end_date)
    
    result = await db.execute(query)
    entries = result.scalars().all()
    
    # Build CSV
    output = io.StringIO()
    output.write("Tarih,Yön,Tutar,KDV,Para Birimi,Not\n")
    
    for entry in entries:
        direction = "Gider" if entry.direction == "expense" else "Gelir"
        output.write(
            f"{entry.entry_date.strftime('%Y-%m-%d')},"
            f"{direction},"
            f"{entry.amount},"
            f"{entry.tax_amount or 0},"
            f"{entry.currency},"
            f'"{entry.notes or ""}"\n'
        )
    
    output.seek(0)
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=muhasebe_export.csv"},
    )


@router.get("/export/xlsx")
async def export_xlsx(
    start_date: datetime | None = None,
    end_date: datetime | None = None,
    db: AsyncSession = Depends(get_db),
):
    """Export ledger entries as XLSX."""
    query = select(LedgerEntry).order_by(LedgerEntry.entry_date.desc())
    
    if start_date:
        query = query.where(LedgerEntry.entry_date >= start_date)
    if end_date:
        query = query.where(LedgerEntry.entry_date <= end_date)
    
    result = await db.execute(query)
    entries = result.scalars().all()
    
    # Build XLSX
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Muhasebe"
    
    # Headers
    headers = ["Tarih", "Yön", "Tutar", "KDV", "Para Birimi", "Not"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)
    
    # Data
    for row_num, entry in enumerate(entries, 2):
        direction = "Gider" if entry.direction == "expense" else "Gelir"
        ws.cell(row=row_num, column=1, value=entry.entry_date.strftime('%Y-%m-%d'))
        ws.cell(row=row_num, column=2, value=direction)
        ws.cell(row=row_num, column=3, value=float(entry.amount))
        ws.cell(row=row_num, column=4, value=float(entry.tax_amount or 0))
        ws.cell(row=row_num, column=5, value=entry.currency)
        ws.cell(row=row_num, column=6, value=entry.notes or "")
    
    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Save to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=muhasebe_export.xlsx"},
    )
